"""
SPARKY - FastAPI Backend
Photovoltaiques Document Management System
"""
import base64
import uuid
import time
from contextlib import asynccontextmanager
from typing import Optional

from fastapi import FastAPI, File, UploadFile, HTTPException, Depends, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import httpx
import structlog

from config import settings
from models import (
    ClassifyRequest, ClassifyResponse,
    ExtractRequest, ExtractResponse,
    CompareRequest, CompareResponse,
    DocumentUploadResponse, DocumentStatus,
    ExportExcelRequest,
    NormalizeLigneRequest, NormalizeLigneResponse,
    ImportDocumentRequest, ImportDocumentResponse,
    StockArticle,
    ComparaisonPrixResponse, ArticlePrixComparaison, PrixFournisseur,
)
from database import get_db
from sqlalchemy.ext.asyncio import AsyncSession
from services import (
    DocumentClassifier, DocumentExtractor, PriceComparator, ProductNormalizer,
    get_comparaison_prix_db, persist_import,
)

log = structlog.get_logger()

# ============================================================
# Singleton services
# ============================================================

classifier  = DocumentClassifier()
extractor   = DocumentExtractor()
comparator  = PriceComparator()
normalizer  = ProductNormalizer()


# ============================================================
# Lifespan
# ============================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    log.info("sparky_startup", env=settings.app_env)
    try:
        from database import init_db
        await init_db()
        log.info("db_tables_ready")
    except Exception as e:
        log.warning("db_init_failed", error=str(e))
    yield
    log.info("sparky_shutdown")


# ============================================================
# App
# ============================================================

app = FastAPI(
    title="SPARKY API",
    description="Document Intelligence for Photovoltaiques - Classification, Extraction, Comparaison",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ============================================================
# ROUTES
# ============================================================

# ============================================================
# AUTH — Login avec PostgreSQL + bcrypt
# ============================================================

from pydantic import BaseModel, EmailStr

class LoginRequest(BaseModel):
    email: EmailStr
    password: str

class RegisterRequest(BaseModel):
    email: EmailStr
    password: str
    nom: str | None = None
    role: str = "user"

@app.post("/auth/login")
async def auth_login(payload: LoginRequest, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import text as _text
    from passlib.hash import bcrypt as _bcrypt

    row = await db.execute(
        _text("SELECT id, email, password_hash, nom, role FROM users WHERE email = :e LIMIT 1"),
        {"e": payload.email},
    )
    user = row.mappings().fetchone()
    if not user or not _bcrypt.verify(payload.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")

    return {
        "user_id": str(user["id"]),
        "email":   user["email"],
        "nom":     user["nom"],
        "role":    user["role"],
    }


@app.post("/auth/register")
async def auth_register(payload: RegisterRequest, db: AsyncSession = Depends(get_db)):
    from sqlalchemy import text as _text
    from passlib.hash import bcrypt as _bcrypt

    user_id = str(uuid.uuid4())
    try:
        await db.execute(
            _text("""
                INSERT INTO users (id, email, password_hash, nom, role)
                VALUES (:id, :email, :hash, :nom, :role)
            """),
            {
                "id": user_id,
                "email": payload.email,
                "hash": _bcrypt.hash(payload.password),
                "nom": payload.nom,
                "role": payload.role,
            },
        )
        await db.commit()
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Utilisateur deja existant ou erreur: {e}")

    return {"user_id": user_id, "email": payload.email, "nom": payload.nom, "role": payload.role}


@app.get("/health")
async def health():
    return {"status": "ok", "service": "sparky-api", "version": "1.0.0"}


# ------------------------------------------------------------
# POST /upload — Upload a document to Supabase Storage
# ------------------------------------------------------------

@app.post("/upload", response_model=DocumentUploadResponse)
async def upload_document(file: UploadFile = File(...)):
    """
    Upload a document (PDF or image) to Supabase Storage.
    Returns the document_id and storage URL.
    """
    allowed_types = {"application/pdf", "image/jpeg", "image/png", "image/tiff"}
    if file.content_type not in allowed_types:
        raise HTTPException(
            status_code=415,
            detail=f"Type de fichier non supporté: {file.content_type}. Formats acceptés: PDF, JPEG, PNG, TIFF",
        )

    document_id = str(uuid.uuid4())
    file_content = await file.read()

    # Upload to Supabase Storage
    storage_url = await _upload_to_supabase(document_id, file.filename, file_content, file.content_type)

    return DocumentUploadResponse(
        document_id=document_id,
        storage_url=storage_url,
        nom_fichier=file.filename or "document",
        statut=DocumentStatus.pending,
    )


# ------------------------------------------------------------
# POST /classify — Classify document type
# ------------------------------------------------------------

@app.post("/classify", response_model=ClassifyResponse)
async def classify_document(request: ClassifyRequest):
    """
    Classify a document into one of 5 categories:
    facture, bon_livraison, bon_commande, avoir, devis.
    """
    file_content = await _resolve_file_content(request.file_base64, request.file_url)
    if not file_content:
        raise HTTPException(status_code=400, detail="Fournir file_base64 ou file_url")

    result = await classifier.classify(
        file_content=file_content,
        mime_type=request.mime_type,
        document_id=request.document_id,
    )

    return result


# ------------------------------------------------------------
# POST /extract — Extract structured data
# ------------------------------------------------------------

@app.post("/extract", response_model=ExtractResponse)
async def extract_document(request: ExtractRequest):
    """
    Extract structured data from a classified document.
    Returns fields specific to document type (facture, devis, etc.)
    """
    file_content = await _resolve_file_content(request.file_base64, request.file_url)
    if not file_content:
        raise HTTPException(status_code=400, detail="Fournir file_base64 ou file_url")

    result = await extractor.extract(
        file_content=file_content,
        mime_type=request.mime_type,
        document_id=request.document_id,
        type_document=request.type_document,
    )

    return result


# ------------------------------------------------------------
# POST /compare — Compare prices across devis
# ------------------------------------------------------------

@app.post("/compare", response_model=CompareResponse)
async def compare_prices(request: CompareRequest):
    """
    Compare prices from multiple devis for the same product reference.
    Highlights the cheapest supplier.
    """
    if len(request.items) < 2:
        raise HTTPException(
            status_code=400,
            detail="Minimum 2 devis nécessaires pour comparer les prix",
        )

    result = await comparator.compare(request)
    return result


# ------------------------------------------------------------
# POST /export/excel — Export articles to Excel
# ------------------------------------------------------------

@app.post("/export/excel")
async def export_excel(request: ExportExcelRequest):
    """Export extracted articles as an Excel file (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl non installé. Lancez: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Articles"

    # ---- Header ----
    headers = [
        "TYPE", "Référence", "Nom", "DESCRIPTION", "Catégorie",
        "Unité de mesure", "Marque", "PRIX ACHAT", "MARGE BENE.",
        "Remise %", "TVA %", "PRIX VENTE", "Quantité totale",
    ]
    col_widths = [10, 20, 35, 35, 15, 15, 12, 14, 12, 10, 8, 14, 14]
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 22

    # ---- Identify best price row (lowest prix_unitaire > 0) ----
    best_idx = -1
    best_val = float("inf")
    for idx, ligne in enumerate(request.lignes):
        v = float(ligne.prix_unitaire) if ligne.prix_unitaire else 0.0
        if v > 0 and v < best_val:
            best_val = v
            best_idx = idx

    # ---- Row fills ----
    even_fill      = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
    odd_fill       = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    best_fill      = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")  # light green
    best_font_bold = Font(size=10, bold=True, color="1B5E20")

    for i, ligne in enumerate(request.lignes, 2):
        prix_achat = float(ligne.prix_unitaire) if ligne.prix_unitaire else None
        tva        = float(ligne.tva_taux)      if ligne.tva_taux      else None
        remise     = float(ligne.remise_pct)    if ligne.remise_pct    else None
        qty        = float(ligne.quantite)      if ligne.quantite      else 1.0

        prix_vente = None
        if prix_achat is not None and tva is not None:
            prix_vente = round(prix_achat * (1 + tva / 100), 3)
        elif prix_achat is not None:
            prix_vente = prix_achat

        is_best = (i - 2) == best_idx  # i starts at 2, idx at 0

        values = [
            "ARTICLE",
            ligne.reference or "",
            ligne.designation,
            ligne.designation,
            "",
            ligne.unite or "pièce(s)",
            ligne.marque or "",
            prix_achat,
            None,
            remise,
            tva,
            prix_vente,
            qty,
        ]
        row_fill = best_fill if is_best else (even_fill if i % 2 == 0 else odd_fill)
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.fill = row_fill
            cell.font = best_font_bold if is_best else Font(size=10)
            if isinstance(val, float) and col >= 8:
                cell.number_format = "#,##0.000"

    # ---- Best price label in last column + 1 ----
    if best_idx >= 0:
        star_col = len(headers) + 1
        ws.cell(row=best_idx + 2, column=star_col, value="★ Meilleur prix").font = Font(
            color="1B5E20", bold=True, size=10
        )

    # ---- Stream response ----
    import io as _io
    buffer = _io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nom = (request.nom_document or "articles").replace(" ", "_")
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nom}_articles.xlsx"},
    )


# ------------------------------------------------------------
# POST /process — All-in-one: upload + classify + extract
# ------------------------------------------------------------

@app.post("/process")
async def process_document(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    One-shot endpoint: upload, classify, and extract a document.
    Returns classification + extracted data in a single call.
    Persists result to PostgreSQL automatically.
    """
    # Read file content once, keep in memory
    file_content = await file.read()
    mime_type = file.content_type or _guess_mime(file.filename or "")
    nom_fichier = file.filename or "document"

    # Upload to storage (dev: mock URL, prod: Supabase)
    import uuid as _uuid
    document_id = str(_uuid.uuid4())
    storage_url = await _upload_to_supabase(document_id, nom_fichier, file_content, mime_type)

    upload_result = DocumentUploadResponse(
        document_id=document_id,
        storage_url=storage_url,
        nom_fichier=nom_fichier,
        statut=DocumentStatus.pending,
    )

    # Classify — pass content directly (no re-download)
    import base64 as _b64
    file_b64 = _b64.b64encode(file_content).decode()

    classify_req = ClassifyRequest(
        document_id=document_id,
        file_base64=file_b64,
        mime_type=mime_type,
    )
    classify_result = await classify_document(classify_req)

    # Extract — pass content directly
    extract_req = ExtractRequest(
        document_id=document_id,
        type_document=classify_result.type_document,
        file_base64=file_b64,
        mime_type=mime_type,
    )
    extract_result = await extract_document(extract_req)

    # Notify n8n after scan (fire-and-forget)
    try:
        donnees = extract_result.donnees or {}
        lignes_raw = donnees.get("lignes", []) or []
        lignes_list = []
        for l in lignes_raw:
            if isinstance(l, dict):
                lignes_list.append({
                    "designation": l.get("designation", "") or "",
                    "reference": l.get("reference_fournisseur", l.get("reference", "")) or "",
                    "quantite": float(l.get("quantite", 1) or 1),
                    "prix_unitaire": float(l.get("prix_unitaire", 0) or 0),
                    "remise": float(l.get("remise_pct", l.get("remise", 0)) or 0),
                    "tva": float(l.get("tva_taux", l.get("tva", 0)) or 0),
                    "montant_ttc": float(l.get("montant_ttc", 0) or 0),
                })
        n8n_payload = {
            "type_document": classify_result.type_document,
            "fournisseur": donnees.get("fournisseur_nom") or donnees.get("fournisseur"),
            "numero_document": donnees.get("numero_facture") or donnees.get("numero_devis") or donnees.get("numero_commande") or donnees.get("numero_bl"),
            "date_document": str(donnees.get("date_facture") or donnees.get("date_devis") or donnees.get("date_commande") or donnees.get("date_livraison") or ""),
            "montant_ht": str(donnees.get("montant_ht", 0) or 0),
            "montant_tva": str(donnees.get("montant_tva", 0) or 0),
            "montant_ttc": str(donnees.get("montant_ttc", 0) or 0),
            "score_confiance": classify_result.score_confiance,
            "modele_ia": classify_result.modele_version,
            "nom_fichier": nom_fichier,
            "mime_type": mime_type,
            "lignes": lignes_list,
        }
        await _notify_n8n(n8n_payload)
    except Exception as e:
        log.warning("n8n_process_notify_failed", error=str(e))

    # Save to PostgreSQL (best-effort — ne bloque pas si DB indisponible)
    try:
        log.info("db_save_start", document_id=document_id, type=classify_result.type_document)
        await _save_document_to_db(
            db=db,
            document_id=document_id,
            nom_fichier=nom_fichier,
            storage_url=storage_url,
            classify_result=classify_result,
            extract_result=extract_result,
        )
        log.info("db_save_ok", document_id=document_id)
    except Exception as e:
        import traceback
        log.error("db_save_failed", error=str(e), trace=traceback.format_exc())

    return {
        "upload": upload_result,
        "classification": classify_result,
        "extraction": extract_result,
    }


# ============================================================
# DOCUMENTS — Lecture depuis PostgreSQL
# ============================================================

@app.get("/documents")
async def get_documents(
    limit: int = 100,
    db: AsyncSession = Depends(get_db),
):
    """
    Retourne la liste des documents scannés stockés en base PostgreSQL.
    Inclut fournisseur, montant TTC, numéro et lignes pour chaque document.
    """
    from sqlalchemy import text as _text

    try:
        rows = await db.execute(_text("""
            SELECT
                d.id            AS document_id,
                d.nom_fichier,
                d.storage_url,
                d.type_document,
                d.statut,
                d.score_confiance,
                d.created_at,
                f.nom           AS fournisseur_nom,
                COALESCE(
                    fa.montant_ttc, dv.montant_ttc,
                    bc.montant_ttc, av.montant_ttc
                )               AS montant_ttc,
                COALESCE(
                    fa.montant_ht, dv.montant_ht,
                    bc.montant_ht, av.montant_ht
                )               AS montant_ht,
                COALESCE(
                    fa.numero_facture, dv.numero_devis,
                    bc.numero_bc, bl.numero_bl, av.numero_avoir
                )               AS numero_document
            FROM documents d
            LEFT JOIN fournisseurs   f  ON f.id  = d.fournisseur_id
            LEFT JOIN factures       fa ON fa.document_id = d.id
            LEFT JOIN devis          dv ON dv.document_id = d.id
            LEFT JOIN bons_commande  bc ON bc.document_id = d.id
            LEFT JOIN bons_livraison bl ON bl.document_id = d.id
            LEFT JOIN avoirs         av ON av.document_id = d.id
            ORDER BY d.created_at DESC
            LIMIT :lim
        """), {"lim": limit})

        documents = []
        for r in rows.mappings():
            doc_id = str(r["document_id"])

            # Charger les lignes selon le type
            lignes = []
            type_doc = r["type_document"]
            if type_doc == "facture":
                lrows = await db.execute(_text("""
                    SELECT lf.designation, lf.reference, lf.quantite, lf.prix_unitaire,
                           lf.remise_pct, lf.tva_taux, lf.montant_ht AS montant_ttc
                    FROM lignes_facture lf
                    JOIN factures fa ON fa.id = lf.facture_id
                    WHERE fa.document_id = :doc
                    ORDER BY lf.position
                """), {"doc": doc_id})
            elif type_doc == "devis":
                lrows = await db.execute(_text("""
                    SELECT ld.designation, ld.reference, ld.quantite, ld.prix_unitaire,
                           ld.remise_pct, NULL AS tva_taux, ld.montant_ht AS montant_ttc
                    FROM lignes_devis ld
                    JOIN devis dv ON dv.id = ld.devis_id
                    WHERE dv.document_id = :doc
                    ORDER BY ld.position
                """), {"doc": doc_id})
            else:
                lrows = None

            if lrows:
                for l in lrows.mappings():
                    lignes.append({
                        "designation":   l["designation"],
                        "reference":     l["reference"],
                        "quantite":      float(l["quantite"] or 0),
                        "prix_unitaire": float(l["prix_unitaire"] or 0),
                        "remise":        float(l["remise_pct"] or 0),
                        "tva":           float(l["tva_taux"] or 0) if l["tva_taux"] else 0,
                        "montant_ttc":   float(l["montant_ttc"] or 0),
                    })

            documents.append({
                "document_id":   doc_id,
                "nom_fichier":   r["nom_fichier"],
                "storage_url":   r["storage_url"],
                "type_document": type_doc,
                "statut":        r["statut"] or "extracted",
                "score_confiance": float(r["score_confiance"]) if r["score_confiance"] else None,
                "fournisseur_nom": r["fournisseur_nom"],
                "montant_ttc":   float(r["montant_ttc"]) if r["montant_ttc"] else None,
                "montant_ht":    float(r["montant_ht"]) if r["montant_ht"] else None,
                "numero_document": r["numero_document"],
                "lignes":        lignes,
                "created_at":    r["created_at"].isoformat() if r["created_at"] else None,
            })

        return documents

    except Exception as e:
        log.warning("db_get_documents_failed", error=str(e))
        return []


@app.delete("/documents/{document_id}")
async def delete_document(document_id: str, db: AsyncSession = Depends(get_db)):
    """Supprime un document. Les tables enfants (factures, devis, lignes_*, ...) sont
    automatiquement nettoyées via les FK ON DELETE CASCADE."""
    from sqlalchemy import text as _text

    try:
        result = await db.execute(_text("DELETE FROM documents WHERE id = :id"), {"id": document_id})
        await db.commit()

        if result.rowcount == 0:
            raise HTTPException(status_code=404, detail="Document introuvable")

        return {"deleted": document_id}

    except HTTPException:
        raise
    except Exception as e:
        await db.rollback()
        log.warning("db_delete_document_failed", document_id=document_id, error=str(e))
        raise HTTPException(status_code=500, detail=f"Suppression impossible: {e}")


# ============================================================
# ARTICLES / STOCK — Catalogue normalisé
# ============================================================

@app.post("/articles/normalize", response_model=NormalizeLigneResponse)
async def normalize_ligne(request: NormalizeLigneRequest):
    """
    Normalise une ligne de document (facture / BL) en article canonique.

    - Si la référence fournisseur est déjà connue → match exact (score 1.0)
    - Si les specs correspondent à un article existant → match par specs (score ≥ 0.80)
    - Sinon → nouvel article créé avec référence_interne générée automatiquement
    """
    return normalizer.normalize_ligne(request)


@app.post("/articles/import", response_model=ImportDocumentResponse)
async def import_document_stock(
    request: ImportDocumentRequest,
    db: AsyncSession = Depends(get_db),
):
    """
    Importe toutes les lignes d'un document dans le catalogue et le stock.

    Pour chaque ligne :
      1. Normalise (détecte article existant par référence, specs, ou similarité texte)
      2. Crée l'article canonique si nécessaire
      3. Persiste dans la DB : articles, articles_fournisseurs, mouvements_stock

    Retourne un résumé : nb_nouveaux, nb_matches, nb_ambigus + détail par ligne.
    """
    # Step 1 — normalize in memory
    result = normalizer.import_document(request)

    # Step 2 — build designation → ligne map for price lookup
    lignes_map = {ligne.designation.strip(): ligne for ligne in request.lignes}

    # Step 3 — persist to DB (best-effort: don't fail the whole import if DB is down)
    try:
        await persist_import(db, request, result, lignes_map)
    except Exception as exc:
        log.warning("articles.persist_failed", error=str(exc))

    return result


@app.get("/articles/stock", response_model=list[StockArticle])
async def get_stock():
    """
    Retourne le stock courant de tous les articles actifs avec leur quantité totale
    (somme de tous les mouvements de tous les fournisseurs).

    Utilise la vue v_stock_articles.  En attendant l'intégration DB, les données
    proviennent du catalogue en mémoire du normalizer.
    """
    articles = []
    for art in normalizer.catalogue._articles.values():
        articles.append(
            StockArticle(
                id=art["id"],
                reference_interne=art["reference_interne"],
                nom_normalise=art["nom_normalise"],
                categorie=art.get("categorie"),
                unite_mesure=art.get("unite_mesure"),
                specifications=art.get("specifications"),
                quantite_totale=_sum_stock(art["id"]),
                nb_fournisseurs=_count_fournisseurs(art["id"]),
                derniere_entree=None,
            )
        )
    return articles


def _sum_stock(article_id: str):
    """Sum movements for a given article from the normalizer's in-memory catalogue."""
    from decimal import Decimal as D
    total = D("0")
    for mvt in normalizer.catalogue._movements.get(article_id, []):
        if mvt["type"] == "sortie":
            total -= D(str(mvt["quantite"]))
        else:
            total += D(str(mvt["quantite"]))
    return total


def _count_fournisseurs(article_id: str) -> int:
    return len({
        fid for (fid, _ref), aid in normalizer.catalogue._supplier_refs.items()
        if aid == article_id
    })


@app.get("/articles/comparaison-prix", response_model=ComparaisonPrixResponse)
async def get_comparaison_prix(
    categorie: Optional[str] = None,
    db: AsyncSession = Depends(get_db),
):
    """
    Tableau comparatif automatique — DB-first, fallback mémoire.

    Pour chaque article stocké en base :
      - liste tous les fournisseurs connus avec leur prix achat
      - identifie le meilleur prix automatiquement
      - calcule l'écart en % entre le moins cher et le plus cher

    Paramètre optionnel : categorie (ex: "Câbles", "Panneaux photovoltaïque")
    """
    # DB-first: try to read from PostgreSQL
    try:
        result = await get_comparaison_prix_db(db, categorie=categorie)
        if result.nb_articles > 0:
            return result
    except Exception as exc:
        log.warning("comparaison_prix.db_fallback", error=str(exc))

    # Fallback: build from in-memory catalogue (dev / no-DB mode)
    from decimal import Decimal as D
    catalogue = normalizer.catalogue

    all_fournisseurs: dict[str, str] = {}
    articles_result: list[ArticlePrixComparaison] = []

    for art in sorted(catalogue._articles.values(),
                      key=lambda a: (a.get("categorie") or "", a.get("nom_normalise") or "")):
        if categorie and art.get("categorie") != categorie:
            continue

        aliases = catalogue._supplier_aliases.get(art["id"], [])
        prix_disponibles = [a for a in aliases if a.get("prix_achat") is not None]
        if not prix_disponibles:
            continue

        for a in prix_disponibles:
            fid = a.get("fournisseur_id", "")
            if fid and fid not in all_fournisseurs:
                all_fournisseurs[fid] = a.get("fournisseur_nom_court", fid)

        meilleur   = min(prix_disponibles, key=lambda a: float(a["prix_achat"]))
        meilleur_p = float(meilleur["prix_achat"])
        pire_p     = max(float(a["prix_achat"]) for a in prix_disponibles)
        ecart_pct  = round((pire_p - meilleur_p) / meilleur_p * 100, 1) \
                     if meilleur_p > 0 and pire_p != meilleur_p else 0.0

        prix_par_fourn = sorted([
            PrixFournisseur(
                fournisseur_id=a.get("fournisseur_id", ""),
                fournisseur_nom=a.get("fournisseur_nom_court", ""),
                nom_fournisseur=a.get("nom_fournisseur", ""),
                reference_fournisseur=a.get("reference_fournisseur"),
                prix_achat=D(str(a["prix_achat"])),
                prix_vente=D(str(a["prix_vente"])) if a.get("prix_vente") is not None else None,
                tva_taux=D(str(a["tva_taux"])) if a.get("tva_taux") is not None else None,
                marque=a.get("marque"),
                est_meilleur_prix=(float(a["prix_achat"]) == meilleur_p),
                surcout_pct=round((float(a["prix_achat"]) - meilleur_p) / meilleur_p * 100, 1)
                            if float(a["prix_achat"]) != meilleur_p else 0.0,
            )
            for a in prix_disponibles
        ], key=lambda p: float(p.prix_achat))

        articles_result.append(ArticlePrixComparaison(
            article_id=art["id"],
            reference_interne=art["reference_interne"],
            nom_normalise=art["nom_normalise"],
            categorie=art.get("categorie"),
            unite_mesure=art.get("unite_mesure"),
            quantite_stock=_sum_stock(art["id"]),
            meilleur_prix_achat=D(str(meilleur_p)),
            meilleur_fournisseur=meilleur.get("fournisseur_nom_court"),
            economie_max_pct=ecart_pct,
            prix_par_fournisseur=prix_par_fourn,
        ))

    return ComparaisonPrixResponse(
        nb_articles=len(articles_result),
        fournisseurs=list(all_fournisseurs.values()),
        articles=articles_result,
    )


# ============================================================
# HELPERS
# ============================================================

async def _resolve_file_content(
    file_base64: Optional[str],
    file_url: Optional[str],
) -> Optional[bytes]:
    if file_base64:
        try:
            return base64.b64decode(file_base64)
        except Exception:
            raise HTTPException(status_code=400, detail="base64 invalide")
    if file_url:
        return await _download_from_url(file_url)
    return None


async def _download_from_url(url: str) -> bytes:
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(url)
        response.raise_for_status()
        return response.content


async def _upload_to_supabase(
    document_id: str,
    filename: str,
    content: bytes,
    content_type: str,
) -> str:
    if not settings.supabase_url:
        # Dev mode: return a mock URL
        return f"http://localhost:8000/documents/{document_id}/{filename}"

    from supabase import create_client
    client = create_client(settings.supabase_url, settings.supabase_key)
    path = f"{document_id}/{filename}"
    client.storage.from_(settings.supabase_bucket).upload(
        path,
        content,
        file_options={"content-type": content_type},
    )
    return client.storage.from_(settings.supabase_bucket).get_public_url(path)


async def _save_document_to_db(
    db: AsyncSession,
    document_id: str,
    nom_fichier: str,
    storage_url: str,
    classify_result,
    extract_result,
) -> None:
    """Persist scanned document + details (facture/devis/…) to PostgreSQL."""
    from sqlalchemy import text as _text
    from datetime import date as _date

    donnees = extract_result.donnees or {}
    _type_raw = classify_result.type_document
    type_doc = _type_raw.value if hasattr(_type_raw, "value") else str(_type_raw)
    score = float(classify_result.score_confiance)
    fournisseur_nom = donnees.get("fournisseur_nom") or donnees.get("fournisseur") or ""

    # 1 — upsert fournisseur
    fourn_id = None
    if fournisseur_nom:
        row = await db.execute(
            _text("SELECT id FROM fournisseurs WHERE nom = :nom LIMIT 1"),
            {"nom": fournisseur_nom},
        )
        existing = row.fetchone()
        if existing:
            fourn_id = str(existing[0])
        else:
            fourn_id = str(uuid.uuid4())
            await db.execute(
                _text("INSERT INTO fournisseurs (id, nom) VALUES (:id, :nom) ON CONFLICT DO NOTHING"),
                {"id": fourn_id, "nom": fournisseur_nom},
            )

    def _parse_date(val):
        """Retourne datetime.date ou None. Gere YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY."""
        from datetime import date as _d, datetime as _dt
        if not val:
            return None
        if isinstance(val, _d):
            return val
        s = str(val).strip()[:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d"):
            try:
                return _dt.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _dec(val):
        try:
            return float(val) if val is not None else None
        except Exception:
            return None

    # 2 — insert document (CAST évite le conflit entre :param et :: de PostgreSQL)
    await db.execute(
        _text("""
            INSERT INTO documents
                (id, nom_fichier, storage_url, type_document, statut, score_confiance, fournisseur_id)
            VALUES
                (:id, :nom, :url,
                 CAST(:type AS document_type),
                 CAST('extracted' AS document_status),
                 :score, :fourn)
            ON CONFLICT (id) DO NOTHING
        """),
        {
            "id": document_id,
            "nom": nom_fichier,
            "url": storage_url,
            "type": type_doc,
            "score": score,
            "fourn": fourn_id,
        },
    )

    # 3 — insert type-specific table
    lignes = donnees.get("lignes", []) or []

    if type_doc == "facture":
        fac_id = str(uuid.uuid4())
        await db.execute(
            _text("""
                INSERT INTO factures
                    (id, document_id, fournisseur_id, numero_facture, date_facture,
                     montant_ht, montant_tva, montant_ttc, devise)
                VALUES
                    (:id, :doc, :fourn, :num,
                     CAST(:dt AS date),
                     :ht, :tva, :ttc, 'TND')
            """),
            {
                "id": fac_id, "doc": document_id, "fourn": fourn_id,
                "num": donnees.get("numero_facture"),
                "dt":  _parse_date(donnees.get("date_facture")),
                "ht":  _dec(donnees.get("montant_ht")),
                "tva": _dec(donnees.get("montant_tva")),
                "ttc": _dec(donnees.get("montant_ttc")),
            },
        )
        for i, l in enumerate(lignes):
            await db.execute(
                _text("""
                    INSERT INTO lignes_facture
                        (id, facture_id, designation, reference, quantite,
                         prix_unitaire, remise_pct, tva_taux, montant_ht, position)
                    VALUES
                        (:id, :fac, :des, :ref, :qty, :pu, :rem, :tva, :mt, :pos)
                """),
                {
                    "id": str(uuid.uuid4()), "fac": fac_id,
                    "des": l.get("designation") or "N/A",
                    "ref": l.get("reference_fournisseur") or l.get("reference"),
                    "qty": _dec(l.get("quantite")),
                    "pu":  _dec(l.get("prix_unitaire")),
                    "rem": _dec(l.get("remise_pct") or l.get("remise")),
                    "tva": _dec(l.get("tva_taux") or l.get("tva")),
                    "mt":  _dec(l.get("montant_ttc") or l.get("montant_ht")),
                    "pos": i,
                },
            )

    elif type_doc == "devis":
        dev_id = str(uuid.uuid4())
        await db.execute(
            _text("""
                INSERT INTO devis
                    (id, document_id, fournisseur_id, numero_devis, date_devis,
                     montant_ht, montant_ttc, devise)
                VALUES
                    (:id, :doc, :fourn, :num, CAST(:dt AS date), :ht, :ttc, 'TND')
            """),
            {
                "id": dev_id, "doc": document_id, "fourn": fourn_id,
                "num": donnees.get("numero_devis"),
                "dt":  _parse_date(donnees.get("date_devis")),
                "ht":  _dec(donnees.get("montant_ht")),
                "ttc": _dec(donnees.get("montant_ttc")),
            },
        )
        for i, l in enumerate(lignes):
            await db.execute(
                _text("""
                    INSERT INTO lignes_devis
                        (id, devis_id, designation, reference, quantite,
                         prix_unitaire, remise_pct, montant_ht, position)
                    VALUES
                        (:id, :dev, :des, :ref, :qty, :pu, :rem, :mt, :pos)
                """),
                {
                    "id": str(uuid.uuid4()), "dev": dev_id,
                    "des": l.get("designation") or "N/A",
                    "ref": l.get("reference_fournisseur") or l.get("reference"),
                    "qty": _dec(l.get("quantite")),
                    "pu":  _dec(l.get("prix_unitaire")),
                    "rem": _dec(l.get("remise_pct") or l.get("remise")),
                    "mt":  _dec(l.get("montant_ttc") or l.get("montant_ht")),
                    "pos": i,
                },
            )

    elif type_doc == "bon_commande":
        await db.execute(
            _text("""
                INSERT INTO bons_commande
                    (id, document_id, fournisseur_id, numero_bc, date_commande,
                     montant_ht, montant_ttc, devise)
                VALUES
                    (:id, :doc, :fourn, :num, CAST(:dt AS date), :ht, :ttc, 'TND')
            """),
            {
                "id": str(uuid.uuid4()), "doc": document_id, "fourn": fourn_id,
                "num": donnees.get("numero_commande") or donnees.get("numero_bc"),
                "dt":  _parse_date(donnees.get("date_commande")),
                "ht":  _dec(donnees.get("montant_ht")),
                "ttc": _dec(donnees.get("montant_ttc")),
            },
        )

    elif type_doc == "bon_livraison":
        await db.execute(
            _text("""
                INSERT INTO bons_livraison
                    (id, document_id, fournisseur_id, numero_bl, date_livraison)
                VALUES
                    (:id, :doc, :fourn, :num, CAST(:dt AS date))
            """),
            {
                "id": str(uuid.uuid4()), "doc": document_id, "fourn": fourn_id,
                "num": donnees.get("numero_bl"),
                "dt":  _parse_date(donnees.get("date_livraison")),
            },
        )

    elif type_doc == "avoir":
        await db.execute(
            _text("""
                INSERT INTO avoirs
                    (id, document_id, fournisseur_id, numero_avoir, date_avoir,
                     montant_ht, montant_ttc, devise)
                VALUES
                    (:id, :doc, :fourn, :num, CAST(:dt AS date), :ht, :ttc, 'TND')
            """),
            {
                "id": str(uuid.uuid4()), "doc": document_id, "fourn": fourn_id,
                "num": donnees.get("numero_avoir"),
                "dt":  _parse_date(donnees.get("date_avoir")),
                "ht":  _dec(donnees.get("montant_ht")),
                "ttc": _dec(donnees.get("montant_ttc")),
            },
        )

    await db.commit()
    log.info("db.document_saved", document_id=document_id, type=str(type_doc))


def _json_safe(obj):
    """Convertit Decimal/Enum/date en types JSON-serialisables."""
    from decimal import Decimal
    from datetime import date, datetime
    from enum import Enum
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_json_safe(v) for v in obj]
    if isinstance(obj, Decimal):
        return float(obj)
    if isinstance(obj, Enum):
        return obj.value
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    return obj


async def _notify_n8n(payload: dict) -> None:
    if not settings.n8n_webhook_url:
        return
    try:
        async with httpx.AsyncClient(timeout=5) as client:
            await client.post(
                settings.n8n_webhook_url,
                json=_json_safe(payload),
                headers={"X-Sparky-Secret": settings.n8n_webhook_secret},
            )
    except Exception as e:
        log.warning("n8n_notify_failed", error=str(e))


def _guess_mime(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "tiff": "image/tiff",
        "tif": "image/tiff",
    }.get(ext, "application/pdf")


# ============================================================
# Entry point
# ============================================================

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(
        "main:app",
        host=settings.app_host,
        port=settings.app_port,
        reload=settings.app_env == "development",
    )
